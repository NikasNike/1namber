
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
import pandas as pd
import csv

wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames
sheet = wb.active
ws = wb['Первый лист']
command = 'nul'
x = 0

while command != 'exit':
    command = input("Введите команду, или 'info': ")

    def save():
        wb.save('example.xlsx')
        print("Заметка сохранена")
        col = csv.writer(open("tt.csv", 
                      'w',  
                      newline="")) 
        for r in ws.rows: 
            col.writerow([cell.value for cell in r])  
        
    
    def info():
        print("'save' - сохранить изминения \
            'read' - прочитать данные \
            'add' - добавить данные \
            'edit' - изменить данные \
            'exit' - выйти без сохранения\
            'delite' - удалить данные")    
    
    def read():
                print("Что бы вывести все введите '0'")
                z1 = input("Введите номер заметки или введите дату в формате 'mm-dd': ")
                if  z1.isdigit() == True:
                    z1 = int(z1)
                    if z1 == 0:
                        for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                            if row[0] == None:
                                return
                            else:
                                print(row)
                    elif isinstance(z1, (int)) == True:
                        for row in ws.iter_rows(min_row=z1, max_row=z1, max_col=4, values_only=True):
                            if row[0] == None:
                                print("такйо заметки нет")
                            else:
                                print(row)
                else:
                     for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                         zz = row[3]
                         if zz[5:10] == z1:
                             print(row)
                         else:
                            print("Такой заметки нет")
                            return
    
    def add():
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
           i = 0
           
        if row[0] is None or isinstance(row[0], (str)) == True:
            izi = 0
        else:
            print(type(row[0]))
            izi = (row[0])
        now = datetime.now()
        today = date.today()
        d0 = input("Введите заголовок заметки: ")
        d = input("Введите тело заметки: ")
        ws.append({1 : (izi + 1), 2 : d0, 3 : d, 4 : "{}  {}:{}".format(today, now.hour, now.minute)})
        print("Заметка успешно сохранена")
        
    def edit():
        d1 = int(input("Ведите какую строку изменить: "))
        d12 = input("Ведите новые данные: ")
        now = datetime.now()
        ws.cell(row=d1, column=3, value=d12)
        ws.cell(row=d1, column=4, value="{}.{}.{}  {}:{}".format(now.day, now.month, now.year, now.hour, now.minute))
        print("Заметка успешно изменена")

    def delite():
        d13 = int(input("введите какую строку удалить: "))
        ws.delete_rows(idx = d13, amount=1)
        print("Заметка успешно удалена")

    try:
        method = {
            'save': save,
            'read': read,
            'add': add,
            'edit': edit,
            'delite': delite,
            'info': info,
            'exit': exit
        }
        method[command.lower()]()

    except LookupError:
        print("Неверная команда")
    else:
        x = 0

print("вы вышли((()))")